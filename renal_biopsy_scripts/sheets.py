"""
Scripts for parsing Worksheet Data.

By default this script saves the mapping between SP# and file names as a JSON.
"""

import sys
import copy
import math
import json
from openpyxl import load_workbook, Workbook
from typing import Union, Dict, NamedTuple
from pathlib import Path
from pandas import DataFrame
from collections import namedtuple

__all__ = ["get_tables", "get_subject_slides_mapping"]

def get_subject_slides_mapping(
    df: DataFrame,
    diagnosis: str,
    )-> Dict[str, Dict[str, str]]:
    """
    Gets the mapping between SP# and slide #.

    Parameters
    ----------
    df: DataFrame
        Pandas DataFrame that maps slides to sheets
    diagnosis: str
        One of ('abmr', 'tcmr', 'other')
    is_rejection: bool
        True if the sheet is for rejection cases
        (needs to be parsed differently)

    Returns
    -------
    mapping: Dict[str, Dict[str, Union[str,int]]]
    """
    is_rejection = diagnosis in ('abmr', 'tcmr')

    if not is_rejection and diagnosis != 'other':
        raise ValueError("diagnosis should be one of 'abmr', 'tcmr', 'other'")

    if is_rejection:
        stains = ["H&E", "PAS", "Trichrome"]

        # The column format is: SP#, H&E, PAS, Trichrome
        # SP# is always there which is why we threshold
        # at 2, to remove rows where there is no mapping
        df_filtered = df.dropna(thresh=2)
        assert df_filtered is not None and not df_filtered.empty
        mapping = {}
        for _, row in df_filtered.iterrows():
            d = { "diagnosis": diagnosis }
            for stain in stains:
                num = float(str(row[stain]))
                if math.isnan(num): continue
                d[stain] = str(int(num))
            mapping[row["SP#"]] = d
        return mapping
    else:
        stain="PAS"
        mapping = {}
        for _, row in df.iterrows():
            d = { "diagnosis": diagnosis, stain: str(row["File Name"])}
            mapping[row["SP#"]] = d
        return mapping

def get_tables(
    rejection_path: Union[Path, str], other_path: Union[Path, str]
) -> NamedTuple:
    """
    Gets spreadsheet data as Pandas DataFrames.

    Parameters
    ----------
    rejection_path: Union[Path, str]
        Path to 'Rejection & Infection Cases.xlsx'
    other_path: Union[Path, str]
        Path to 'Other Cases.xlsx'

    Returns
    -------
    dataframes: NamedTuple
        A tuple containing all data as pandas DataFrames.
    """

    wb_rejection = load_workbook(rejection_path)
    wb_other = load_workbook(other_path)

    Tables = namedtuple(
        "Tables", field_names=["tcmr", "abmr", "tcmr_slides", "abmr_slides", "other"]
    )

    return Tables(
        tcmr=parse_tcmr_sheet(wb_rejection["TCMR"]),
        abmr=parse_abmr_sheet(wb_rejection["ABMR"]),
        tcmr_slides=parse_slides_sheet(wb_rejection["TCMR Slides"]),
        abmr_slides=parse_slides_sheet(wb_rejection["ABMR Slides"]),
        other=parse_other_workbook(wb_other),
    )


def parse_other_workbook(wb: Workbook):
    sheet = wb["Sheet1"]

    columns = [
        "File Name",
        "SP#",
        "Stain",
        "Date",
        "Bx Reason",
        "Time Post-Tx",
        "EMPTY",
        "3mo Date",
        "3mo Scr",
        "6mo Date",
        "6mo Scr",
        "1yr Date",
        "1yr SCr",
        "2yr Date",
        "2yr SCr",
        "3yr Date",
        "3yr SCr",
        "EMPTY",
        "Latest Date",
        "Latest SCr",
        "EMPTY",
        "Date of Tx",
        "DOB",
        "Gender",
        "Race",
    ]

    top_left, bottom_right = "A2", "Y186"  # TODO: what to do about the missing columns?

    # # TODO : parse date using this module
    # deal_with_it = {
    #     "DOB": lambda x: datetime.strptime(x, "")
    # }

    table = []
    for sheet_row in sheet[top_left:bottom_right]:
        d = {k: (v.value if v.value else "") for k, v in zip(columns, sheet_row)}
        table.append(d)

    df = DataFrame.from_dict(table)
    del df["EMPTY"]
    return df


def parse_slides_sheet(sheet):
    columns = [
        "SP# pre",
        "SP# post",
        "H&E",
        "PAS",
        "Trichrome",
    ]
    table_columns = [
        "SP#",
        "H&E",
        "PAS",
        "Trichrome",
    ]

    deal_with_it = {
        "SP#": lambda d: "".join((d["SP# pre"], str(d["SP# post"]))),
    }

    table = []
    for row in sheet.iter_rows(min_row=2, max_col=5):
        if row[0].value is None:
            break
        d = {k: v.value for k, v in zip(columns, row)}
        table_row = {k: d[k] if k in d else deal_with_it[k](d) for k in table_columns}
        table.append(table_row)

    return DataFrame(table)


def parse_abmr_sheet(sheet):

    columns_2020 = [
        "Status",
        "Pending",
        "MD",
        "Date",
        "SP# pre",
        "SP# post",
        "Bx Reason",
        "Time Post-Tx pre",
        "Time Post-Tx post",
        "Primary Dz",
        "Interesting findings",
        "Primary Dx",
        "Secondary Dx pre",
        "Secondary Dx post",
        "Acute ABMR",
        "Chronic Rejection",
        "Other Changes #1",
        "Other Changes #2",
        "Specify",
        "Glom #",
        "GS",
        "SS",
        "IF (%)",
        "TA (%)",
        "g",
        "cg",
        "mm",
        "ci",
        "ct",
        "i",
        "ti",
        "t",
        "ptc",
        "ah",
        "aah",
        "cv",
        "v",
        "C4d",
        "i-IFTA",
        "t-IFTA",
    ]

    columns_2019_2018 = [
        "Status",
        "Pending",
        "MD",
        "Date",
        "SP# pre",
        "SP# post",
        "Bx Reason",
        "Time Post-Tx pre",
        "Time Post-Tx post",
        "Primary Dz",
        "Interesting findings",
        "Primary Dx",
        "Secondary Dx",
        "Acute ABMR",
        "Chronic Rejection",
        "Other Changes #1",
        "Other Changes #2",
        "Specify",
        "Glom #",
        "GS",
        "SS",
        "IF (%)",
        "TA (%)",
        "g",
        "cg",
        "mm",
        "ci",
        "ct",
        "i",
        "ti",
        "t",
        "ptc",
        "ah",
        "aah",
        "cv",
        "v",
        "C4d",
        "i-IFTA",
        "t-IFTA",
    ]

    table_columns = [
        "SP#",
        "Bx Reason",
        "Time Post-Tx",
        "Primary Dz",
        "Interesting findings",
        "Primary Dx",
        "Secondary Dx",
        "Acute ABMR",
        "Chronic Rejection",
        "Other Changes #1",
        "Other Changes #2",
        "Specify",
        "Glom #",
        "GS",
        "SS",
        "IF (%)",
        "TA (%)",
        "g",
        "cg",
        "mm",
        "ci",
        "ct",
        "i",
        "ti",
        "t",
        "ptc",
        "ah",
        "aah",
        "cv",
        "v",
        "C4d",
        "i-IFTA",
        "t-IFTA",
    ]

    years = ["2020", "2019", "2018"]
    year_positions = {
        "2020": ("A3", "AN16"),
        "2019": ("A20", "AM36"),
        "2018": ("A40", "AM46"),
    }

    deal_with_it = {
        "SP#": lambda d: "".join((d["SP# pre"], str(d["SP# post"]))),
        "Secondary Dx": lambda d:  " ".join(
            (d["Secondary Dx pre"], str(d["Secondary Dx post"]))),
        "Time Post-Tx": lambda d: " ".join(
            (str(d["Time Post-Tx pre"]), str(d["Time Post-Tx post"]).lower())
        ),
        "Unnamed_col": lambda _: "",
    }

    table = []
    for year in years:
        top_left, bottom_right = year_positions[year]
        columns = columns_2020 if year == "2020" else columns_2019_2018
        for sheet_row in sheet[top_left:bottom_right]:
            d = {k: (v.value if v.value else "") for k, v in zip(columns, sheet_row)}
            table_row = {
                k: d[k] if k in d else deal_with_it[k](d) for k in table_columns
            }
            table_row["Year"] = year
            if "RIGHT(" in str(table_row["SP#"]):
                spnum = int(str(table_row["SP#"].split("\"")[-1]))
                table_row["SP#"] = "SP-"+str(spnum)
            if "SP" not in str(table_row["SP#"]):
                table_row["SP#"] = "SP-"+str(table_row["SP#"])
            table_row["Year"] = year
            table.append(table_row)

    return DataFrame(table)


def parse_tcmr_sheet(sheet):
    years = ["2020", "2019", "2018", "2017", "2016"]
    cols_2020_2018 = [
        "Status",
        "Pending",
        "MD",
        "Date",
        "SP# prefix",
        "SP# postfix",
        "Bx Reason",
        "Time Post-Tx num",
        "Time Post-Tx units",
        "Primary Dz",
        "Interesting findings",
        "Primary Dx",
        "Secondary Dx",
        "Acute TCMR",
    ]
    cols_2017_2016 = [
        "Date",
        "SP#",
        "Status",
        "Pending",
        "Specimen Type",
        "Bx Reason",
        "Time Post-Tx",
        "Primary Dz",
        "Cr (baseline)",
        "Cr (current)",
        "Proteinuria",
        "Clinical Hx",
        "Interesting findings",
        "Primary Dx",
        "Secondary Dx",
    ]
    all_unique_cols = [
        "Acute TCMR",
        "Bx Reason",
        "Clinical Hx",
        "Cr (baseline)",
        "Cr (current)",
        "Date",
        "Interesting findings",
        "MD",
        "Pending",
        "Primary Dx",
        "Primary Dz",
        "Proteinuria",
        "SP#",
        "Secondary Dx",
        "Specimen Type",
        "Status",
        "Time Post-Tx",
    ]
    year_positions = {
        "2020": ("A3", "N11"),
        "2019": ("A15", "N29"),
        "2018": ("A33", "N47"),
        "2017": ("A52", "P73"),
        "2016": ("A78", "P83"),
    }

    # Dictionary with mapped behaviour for missing or weird columns.
    # This keeps the main loop short and all of the adjustments that have to be
    # made in one single place.
    deal_with_it = {
        #########################
        # Merge columns
        #########################
        "SP#": lambda d: "".join(
            [d["SP# prefix"], str(d["SP# postfix"])]
        ),  # TODO: is the SP- prefix needed?
        "Time Post-Tx": lambda d: " ".join(
            [str(i) for i in (d["Time Post-Tx num"], d["Time Post-Tx units"].lower())]
        ),
        #########################
        # All columns missing from earlier years
        #########################
        "Acute TCMR": lambda _: "",
        "MD": lambda _: "",
        #########################
        # All columns missing from the more recent years
        #########################
        "Clinical Hx": lambda _: "",
        "Cr (baseline)": lambda _: "",
        "Cr (current)": lambda _: "",
        "Proteinuria": lambda _: "",
        "Specimen Type": lambda _: "",
    }
    table = []
    for year in years:
        top_left, bottom_right = year_positions[year]
        columns = cols_2020_2018 if year[-1] in "098" else cols_2017_2016
        for sheet_row in sheet[top_left:bottom_right]:
            d = {k: (v.value if v.value else "") for k, v in zip(columns, sheet_row)}
            table_row = {
                k: d[k] if k in d else deal_with_it[k](d) for k in all_unique_cols
            }
            if "RIGHT(" in str(table_row["SP#"]):
                spnum = int(str(table_row["SP#"].split("\"")[-1]))
                table_row["SP#"] = "SP-"+str(spnum)

            if "SP" not in str(table_row["SP#"]):
                table_row["SP#"] = "SP-"+str(table_row["SP#"])
            table_row["Year"] = year
            table.append(table_row)
    return DataFrame(table)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <path_to_rejection_cases> <path_to_other_cases>")
        exit(0)
    rejection_path = sys.argv[1]
    other_path = sys.argv[2]
    tables = get_tables(rejection_path, other_path)
    tcmr_slides, abmr_slides, non_rejection_slides = [tables[i+2] for i in range(3)]
    tcmr_map, abmr_map, nr_map = (get_subject_slides_mapping(tcmr_slides, "tcmr"),
                                  get_subject_slides_mapping(abmr_slides, "abmr"),
                                  get_subject_slides_mapping(non_rejection_slides, "other"))
    combined_mapping = {**tcmr_map, **abmr_map, **nr_map}
    with open ("subject2file.json", "w") as f: json.dump(combined_mapping, f)
